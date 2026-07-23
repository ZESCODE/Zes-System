"""Read/write CSV and Excel files."""
import csv, io, os
from typing import Dict, Any, List, Optional
import aiofiles
import openpyxl


class SpreadsheetSkill:
    async def read_csv(self, path: str, delimiter: str = ",") -> Dict[str, Any]:
        try:
            async with aiofiles.open(path, 'r', encoding='utf-8') as f:
                content = await f.read()
            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            return {"success": True, "rows": [row for row in reader]}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def write_csv(self, path: str, rows: List[List[str]], delimiter: str = ",") -> Dict[str, Any]:
        try:
            output = io.StringIO()
            writer = csv.writer(output, delimiter=delimiter)
            writer.writerows(rows)
            async with aiofiles.open(path, 'w', encoding='utf-8') as f:
                await f.write(output.getvalue())
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def read_excel(self, path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active
            return {"success": True, "data": [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def write_excel(self, path: str, data: List[List[Any]], sheet_name: str = "Sheet1") -> Dict[str, Any]:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            for row in data:
                ws.append(row)
            wb.save(path)
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}
